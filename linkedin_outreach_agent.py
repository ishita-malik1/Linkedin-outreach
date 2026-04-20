"""
LinkedIn outreach agent: reads job tracker + resume, searches the web (Serper),
fetches pages, and uses Gemini with tool calling to draft personalized messages.

Requires: GEMINI_API_KEY and SERPER_API_KEY in .env (see .env.example).
"""

# Do not use `from __future__ import annotations` here: google-generativeai builds Pydantic
# schemas from tool function signatures, and postponed (string) annotations break Optional/Union.

import argparse
import json
import os
import random
import re
import sys
import time
from pathlib import Path
from typing import Any, Optional

import httpx
from google.api_core.exceptions import ResourceExhausted
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from pypdf import PdfReader

load_dotenv()

try:
    import google.generativeai as genai
except ImportError as e:
    raise SystemExit(
        "Missing dependency google-generativeai. Run: pip install -r requirements.txt"
    ) from e

# --- Paths & config -----------------------------------------------------------------

DEFAULT_EXCEL = Path("data/jobs.xlsx")
DEFAULT_RESUME = Path("data/resume.pdf")
# Default: 2.5 Flash-Lite is usually easier on free-tier RPM than 2.0 Flash; override via GEMINI_MODEL.
DEFAULT_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash-lite")

SERPER_URL = "https://google.serper.dev/search"
MAX_FETCH_CHARS = 14_000

_resume_cache: str | None = None
_outcomes: list[dict[str, Any]] = []


# --- Spreadsheet helpers -------------------------------------------------------------

def _norm(s: str) -> str:
    return " ".join(s.strip().split()).lower()


def _find_header_map(ws) -> dict[str, int]:
    """Map canonical keys to 1-based column indexes from first row."""
    header_row = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        header_row.append("" if v is None else str(v).strip())

    aliases = {
        "company": {"company", "company name", "employer"},
        "title": {"job title", "title", "role", "position"},
        "location": {"location", "city"},
        "status": {"application status", "status", "stage"},
    }
    col_map: dict[str, int] = {}
    lowered = [_norm(h) for h in header_row]

    for key, allowed in aliases.items():
        for idx, h in enumerate(lowered, start=1):
            if h in allowed:
                col_map[key] = idx
                break
        if key not in col_map:
            for idx, raw in enumerate(header_row, start=1):
                if raw and _norm(raw) in allowed:
                    col_map[key] = idx
                    break

    missing = [k for k in aliases if k not in col_map]
    if missing:
        raise ValueError(
            f"Could not find columns for: {missing}. Headers seen: {header_row}"
        )
    return col_map


def _row_dict(ws, row_idx: int, col_map: dict[str, int]) -> dict[str, str]:
    def cell(col: int) -> str:
        v = ws.cell(row=row_idx, column=col).value
        return "" if v is None else str(v).strip()

    return {
        "company": cell(col_map["company"]),
        "job_title": cell(col_map["title"]),
        "location": cell(col_map["location"]),
        "status": cell(col_map["status"]),
    }


def get_last_non_rejected_rows(path: Path, limit: int) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        col_map = _find_header_map(ws)
        picked: list[dict[str, str]] = []
        for r in range(ws.max_row, 1, -1):
            row = _row_dict(ws, r, col_map)
            if not any(row.values()):
                continue
            st = _norm(row["status"])
            if st == "rejected":
                continue
            picked.append(row)
            if len(picked) >= limit:
                break
        return list(reversed(picked))
    finally:
        wb.close()


# --- Resume -------------------------------------------------------------------------

def extract_resume_text(path: Path) -> str:
    reader = PdfReader(str(path))
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        parts.append(t)
    return "\n".join(parts).strip()


# --- HTTP tools ---------------------------------------------------------------------

def serper_search(query: str, api_key: str) -> dict[str, Any]:
    payload = {"q": query, "num": 10}
    r = httpx.post(
        SERPER_URL,
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json=payload,
        timeout=30.0,
    )
    r.raise_for_status()
    return r.json()


def fetch_page_text(url: str) -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; JobSearchBot/1.0; +https://example.invalid)"
        )
    }
    r = httpx.get(url, headers=headers, follow_redirects=True, timeout=25.0)
    r.raise_for_status()
    ctype = r.headers.get("content-type", "")
    if "text/html" not in ctype and "application/xhtml" not in ctype:
        return f"[Non-HTML content skipped: {ctype}]\n" + r.text[:2000]

    soup = BeautifulSoup(r.text, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(separator="\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = "\n".join(line.strip() for line in text.splitlines())
    text = text.strip()
    if len(text) > MAX_FETCH_CHARS:
        text = text[:MAX_FETCH_CHARS] + "\n...[truncated]"
    return text


# ====================================================================================
# Gemini tools (docstrings are exposed to the model)
# ====================================================================================

_excel_path_arg: Path = DEFAULT_EXCEL
_resume_path_arg: Path = DEFAULT_RESUME


def get_pending_applications(limit: int = 5) -> str:
    """Return the last `limit` spreadsheet rows where application status is not \"rejected\".

    Each row includes company name, job title, location, and status. Rows are ordered
    oldest-first among the selected tail (same order as typical tracker workflow).
    """
    rows = get_last_non_rejected_rows(_excel_path_arg, limit)
    return json.dumps(rows, indent=2)


def read_resume_text() -> str:
    """Return full plain text extracted from the user's resume PDF."""
    global _resume_cache
    if _resume_cache is None:
        if not _resume_path_arg.is_file():
            return f"[ERROR] Resume file not found at {_resume_path_arg}"
        _resume_cache = extract_resume_text(_resume_path_arg)
    if not _resume_cache.strip():
        return "[WARNING] Resume text is empty — check PDF extractability."
    return _resume_cache


def web_search(query: str) -> str:
    """Search the public web via Serper (Google results). Use precise queries: job postings,
    company career pages, and LinkedIn profiles (site:linkedin.com/in ...).
    """
    key = os.environ.get("SERPER_API_KEY", "").strip()
    if not key:
        return json.dumps({"error": "SERPER_API_KEY missing in environment"})
    try:
        data = serper_search(query, key)
    except Exception as e:
        return json.dumps({"error": str(e)})
    organic = data.get("organic") or []
    slim = []
    for item in organic[:10]:
        slim.append(
            {
                "title": item.get("title"),
                "link": item.get("link"),
                "snippet": item.get("snippet"),
            }
        )
    return json.dumps({"organic": slim, "knowledge_graph": data.get("knowledgeGraph")}, indent=2)


def fetch_webpage(url: str) -> str:
    """Fetch a public HTTP(S) URL and return visible text (HTML stripped). Many sites block
    bots; if text is empty or blocked, rely on search snippets instead.
    """
    try:
        return fetch_page_text(url)
    except Exception as e:
        return f"[fetch error] {e}"


def submit_role_outcome(
    company_name: str,
    job_title: str,
    contact_name: Optional[str],
    contact_title: Optional[str],
    linkedin_profile_url: Optional[str],
    skipped: bool,
    skip_reason: Optional[str],
    connection_note_under_300_chars: str,
    inmail_100_to_150_words: str,
) -> str:
    """Submit final structured results for ONE role after searching and drafting messages.

    If you cannot identify a plausible hiring manager or recruiter LinkedIn profile after a few
    searches, set skipped=true and explain briefly in skip_reason (leave URLs/name empty).

    connection_note_under_300_chars: connection request note, hard max 300 characters.
    inmail_100_to_150_words: longer message body (aim 100–150 words), human and specific.
    """
    note = connection_note_under_300_chars or ""
    if len(note) > 300:
        note = note[:297] + "..."

    wc = len(inmail_100_to_150_words.split())
    warn = ""
    if not skipped and (wc < 90 or wc > 170):
        warn = f" [word_count={wc}, target 100-150]"

    _outcomes.append(
        {
            "company_name": company_name,
            "job_title": job_title,
            "contact_name": contact_name,
            "contact_title": contact_title,
            "linkedin_profile_url": linkedin_profile_url,
            "skipped": skipped,
            "skip_reason": skip_reason,
            "connection_note": note,
            "inmail": inmail_100_to_150_words.strip(),
            "word_count_warning": warn,
        }
    )
    return "Recorded outcome for " + company_name + warn


# --- System prompt ------------------------------------------------------------------

SYSTEM_INSTRUCTION = """You are an orchestrator for job-search LinkedIn outreach.

Goal: For each pending application row, find a specific job posting (or strong match), then
identify ONE best contact at that company on LinkedIn: prefer hiring manager for the team/role,
otherwise a recruiter supporting that function/location.

Use tools freely: web_search with targeted queries, fetch_webpage on promising URLs for extra
detail. LinkedIn often blocks scraping — it is OK to rely on accurate Google/Serper snippets that
include names/titles and linkedin.com/in URLs.

Read the resume once with read_resume_text early. Pull 1–2 concrete alignment points between the
resume and THIS role (tools, domain, measurable outcomes).

Writing rules for messages:
- Sound like a thoughtful human; vary sentence openings.
- No filler ("I hope this finds you well"), no buzzword soup.
- Reference something concrete from the posting or the person's background when possible.
- Produce submit_role_outcome exactly once per role after you finish that role.

Search discipline:
- If you cannot find a plausible hiring manager or recruiter after several distinct searches,
  skip: skipped=true with a short skip_reason.

Process every row returned by get_pending_applications (typically up to 5). Order is your choice.
"""


USER_TASK_TEMPLATE = """Run the outreach workflow end-to-end.

1) Call get_pending_applications() to list roles (last 5 non-rejected rows).
2) Call read_resume_text() once you start drafting (or earlier).
3) For each role: search → optionally fetch pages → choose contact → submit_role_outcome.

When everything is submitted, reply with a one-line summary only."""


# --- Gemini run ---------------------------------------------------------------------

_TOOL_REGISTRY: dict[str, Any] = {
    "get_pending_applications": get_pending_applications,
    "read_resume_text": read_resume_text,
    "web_search": web_search,
    "fetch_webpage": fetch_webpage,
    "submit_role_outcome": submit_role_outcome,
}


def _parse_retry_after_seconds(exc: ResourceExhausted) -> float | None:
    err = str(exc)
    m = re.search(r"retry in ([\d.]+)\s*s", err, flags=re.I)
    if m:
        return float(m.group(1))
    return None


def _gemini_send_with_retry(chat: Any, contents: Any, *, max_attempts: int = 10) -> Any:
    """Retry on 429 ResourceExhausted (free-tier RPM / burst limits)."""
    base_delay = 5.0
    for attempt in range(max_attempts):
        try:
            return chat.send_message(contents)
        except ResourceExhausted as e:
            if attempt >= max_attempts - 1:
                raise
            parsed = _parse_retry_after_seconds(e)
            wait = parsed if parsed is not None else base_delay
            wait = wait + random.uniform(0.5, 2.5)
            print(
                f"Gemini rate limit (429). Sleeping {wait:.1f}s then retrying "
                f"({attempt + 1}/{max_attempts})…",
                file=sys.stderr,
            )
            time.sleep(wait)
            base_delay = min(base_delay * 1.5, 120.0)
    raise RuntimeError("unreachable")


def _dispatch_tool_call(name: str, args: dict[str, Any]) -> str:
    fn = _TOOL_REGISTRY.get(name)
    if not fn:
        return json.dumps({"error": f"unknown tool {name}"})
    try:
        return fn(**args)
    except TypeError:
        return json.dumps({"error": f"bad arguments for {name}", "args": args})
    except Exception as e:
        return json.dumps({"error": str(e)})


def run_gemini_agent() -> None:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        print("Set GEMINI_API_KEY in .env", file=sys.stderr)
        sys.exit(1)
    if not os.environ.get("SERPER_API_KEY", "").strip():
        print(
            "Warning: SERPER_API_KEY is empty — web_search will fail until you add it to .env.",
            file=sys.stderr,
        )

    genai.configure(api_key=api_key)

    print(f"Using Gemini model: {DEFAULT_MODEL}", file=sys.stderr)

    tools = [
        get_pending_applications,
        read_resume_text,
        web_search,
        fetch_webpage,
        submit_role_outcome,
    ]

    model = genai.GenerativeModel(
        model_name=DEFAULT_MODEL,
        system_instruction=SYSTEM_INSTRUCTION,
        tools=tools,
    )

    # Manual tool loop only (single execution path; avoids double-invoking tools with auto-Fn).
    chat = model.start_chat()
    response = _gemini_send_with_retry(chat, USER_TASK_TEMPLATE)

    max_turns = 80
    for _ in range(max_turns):
        if not response.candidates:
            break
        parts = list(response.candidates[0].content.parts)
        fc_list = [p for p in parts if getattr(p, "function_call", None)]
        if not fc_list:
            break

        fr_parts: list[Any] = []
        for p in fc_list:
            fc = p.function_call
            args: dict[str, Any] = {}
            if fc.args:
                for k, v in fc.args.items():
                    args[str(k)] = v
            out = _dispatch_tool_call(fc.name, args)
            fr = genai.protos.FunctionResponse(
                name=fc.name,
                response={"result": out},
            )
            fr_parts.append(genai.protos.Part(function_response=fr))

        response = _gemini_send_with_retry(chat, fr_parts)

    if response.text:
        print(response.text.strip())


def print_outcome_cards() -> None:
    if not _outcomes:
        print("\n(No structured outcomes recorded — did the model call submit_role_outcome?)")
        return

    for i, o in enumerate(_outcomes, start=1):
        sep = "=" * 72
        print(f"\n{sep}\n ROLE {i}/{len(_outcomes)}  {o['company_name']} — {o['job_title']}\n{sep}")
        if o.get("skipped"):
            print(" Status: SKIPPED")
            print(f" Reason: {o.get('skip_reason') or '—'}")
            continue

        print(f" Contact: {o.get('contact_name') or '—'}")
        print(f" Title:   {o.get('contact_title') or '—'}")
        print(f" LinkedIn: {o.get('linkedin_profile_url') or '—'}")

        print("\n--- Connection note (≤300 chars) ---")
        print(o.get("connection_note") or "")
        print(f"[length: {len(o.get('connection_note') or '')} chars]")

        print("\n--- InMail draft (100–150 words target) ---")
        print(o.get("inmail") or "")
        wc = len((o.get("inmail") or "").split())
        print(f"[words: {wc}]{o.get('word_count_warning', '')}")


# --- Sample data --------------------------------------------------------------------

def write_sample_excel(path: Path) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "applications"
    ws.append(["company name", "job title", "location", "application status"])
    ws.append(["Acme Labs", "Senior Backend Engineer", "Remote / US", "applied"])
    ws.append(["Globex", "Platform Engineer", "Austin, TX", "screening"])
    ws.append(["Initech", "DevOps Engineer", "Chicago, IL", "rejected"])
    ws.append(["Umbrella Health", "ML Engineer", "Boston, MA", "applied"])
    ws.append(["Stark Industries", "Staff Software Engineer", "San Francisco, CA", "applied"])
    ws.append(["Wayne Enterprises", "Security Engineer", "New York, NY", "offer"])
    wb.save(path)


def main() -> None:
    global _excel_path_arg, _resume_path_arg

    p = argparse.ArgumentParser(description="Gemini LinkedIn outreach agent")
    p.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    p.add_argument("--resume", type=Path, default=DEFAULT_RESUME)
    p.add_argument(
        "--init-sample-excel",
        action="store_true",
        help="Write a sample data/jobs.xlsx (overwrites if present)",
    )
    args = p.parse_args()

    if args.init_sample_excel:
        write_sample_excel(args.excel.resolve())
        print(f"Wrote sample tracker: {args.excel.resolve()}")
        return

    _excel_path_arg = args.excel.resolve()
    _resume_path_arg = args.resume.resolve()

    if not _excel_path_arg.is_file():
        print(
            f"Excel not found: {_excel_path_arg}. Create it or run with --init-sample-excel.",
            file=sys.stderr,
        )
        sys.exit(1)
    if not _resume_path_arg.is_file():
        print(
            f"Resume PDF not found: {_resume_path_arg}. Add data/resume.pdf.",
            file=sys.stderr,
        )
        sys.exit(1)

    run_gemini_agent()
    print_outcome_cards()


if __name__ == "__main__":
    main()
